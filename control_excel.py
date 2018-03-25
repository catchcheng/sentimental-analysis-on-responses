##control the excel program by writing codes in Python


from openpyxl import load_workbook, Workbook
from openpyxl.cell import Cell
import xlrd

response = load_workbook('brick.xlsx')
response = response.get_sheet_by_name('Sheet1')
response1 = xlrd.open_workbook('brick.xlsx')
response1 = response1.sheet_by_index(0)
rating = load_workbook('ratings-brick-clever.xlsx')
rating = rating.get_sheet_by_name('Sheet1')

wb = Workbook()
ws = wb.active


#ws['A6'] = response1.cell(row,column).value

responsedict = {}
ratingdict = {}
for i in range(30):
    responsedict['list{}'.format(i)] = []
    ratingdict['list{}'.format(i)] = []

for i in range(30):
    for row in response.get_squared_range(min_col=i+1,min_row=1,max_col=i+1,max_row=1379):
        for cell in row:
            responsedict['list{}'.format(i)].append(cell.value)

for i in range(30):
    for row in rating.get_squared_range(min_col=i+1,min_row=1,max_col=i+1,max_row=1379):
        for cell in row:
            ratingdict['list{}'.format(i)].append(cell.value)

for j in range(1,60,2):
    for i in range(1,1380):
        ws.cell(row=i,column=j).value = responsedict['list{}'.format(int((j+1)/2)-1)][i-1]
        ws.cell(row=i,column=j+1).value = ratingdict['list{}'.format(int((j+1)/2)-1)][i-1]
        
            
wb.save('brick-clever.xlsx')
